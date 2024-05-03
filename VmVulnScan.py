import logging
import requests
from azure.identity import DefaultAzureCredential
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)

# Define your authentication parameters
resource = 'https://api.securitycenter.microsoft.com'

# Get access token using DefaultAzureCredential
credential = DefaultAzureCredential()

machine_url = 'https://api.securitycenter.microsoft.com/api/machines'
vulnerabilities_url = 'https://api.securitycenter.microsoft.com/api/vulnerabilities/machinesVulnerabilities'

def fetch_machine_display_names():
    try:
        access_token = credential.get_token(resource).token

        # Set headers with authorization token
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }

        # Make a GET request to fetch machine display names
        response = requests.get(machine_url, headers=headers)
        response.raise_for_status()
        
        # Logging the response content
        logging.info("Response from API: %s", response.content.decode('utf-8'))

        machine_data = response.json()

        # Check if the response contains any data
        if 'value' in machine_data and isinstance(machine_data['value'], list):
            # Extract machine display names from the response
            machine_display_names = {}
            for machine in machine_data['value']:
                machine_id = machine['id']
                computer_dns_name = machine['computerDnsName']
                # Extract machine name before the first dot
                machine_name = computer_dns_name.split('.')[0]
                machine_display_names[machine_id] = machine_name
            return machine_display_names
        else:
            logging.warning("No machine actions found in the response")
            return {}
    except Exception as e:
        logging.error("Error fetching machine display names: %s", e)
        return {}

def fetch_vulnerabilities():
    try:
        access_token = credential.get_token(resource).token

        # Set headers with authorization token
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }

        # Make request to get vulnerabilities
        response = requests.get(vulnerabilities_url, headers=headers)
        response.raise_for_status()
        vulnerabilities_data = response.json().get('value', [])
        logging.info("Vulnerabilities data retrieved successfully")
        return vulnerabilities_data
    except Exception as e:
        logging.error(f"Failed to retrieve vulnerabilities data: {e}")
        return []

# Fetch machine display names
display_name_mapping = fetch_machine_display_names()

# Fetch vulnerabilities
vulnerabilities_data = fetch_vulnerabilities()

# Group vulnerabilities by machine
machine_vulnerabilities = {}
for vulnerability in vulnerabilities_data:
    machine_id = vulnerability.get('machineId')
    product_name = vulnerability.get('productName')
    product_vendor = vulnerability.get('productVendor')
    product_version = vulnerability.get('productVersion')
    severity = vulnerability.get('severity')
    cve_id = vulnerability.get('cveId')
    if machine_id not in machine_vulnerabilities:
        machine_vulnerabilities[machine_id] = {}
    if product_version not in machine_vulnerabilities[machine_id]:
        machine_vulnerabilities[machine_id][product_version] = {'productName': product_name, 'productVendor': product_vendor, 'severity': severity, 'cves': set()}
    machine_vulnerabilities[machine_id][product_version]['cves'].add(cve_id)

# Create a new workbook
workbook = Workbook()

# Get the active worksheet
worksheet = workbook.active
worksheet.title = "Vulnerabilities"

# Write header row
header = ["Machine", "Product Name", "Product Vendor", "Product Version", "Severity", "CVEs"]
worksheet.append(header)

# Keep track of written machine names
written_machines = set()

# Write data rows
for machine_id, versions in machine_vulnerabilities.items():
    machine_name = display_name_mapping.get(machine_id, 'Unknown')
    # Write machine name once if not already written
    if machine_name not in written_machines:
        bold_font = Font(bold=True)
        cell = worksheet.cell(row=worksheet.max_row + 1, column=1)
        cell.value = machine_name
        cell.font = bold_font
        written_machines.add(machine_name)
    for product_version, details in versions.items():
        cves = ', '.join(details['cves'])
        worksheet.append(['', details['productName'], details['productVendor'], product_version, details['severity'], cves])

#Date and path for the saved file
current_date = datetime.now().strftime("%Y.%m.%d")
output_directory = r'C:\Users\BrettBauer\Vulnerability Scanning\2024\VMscans'

# Save the workbook
excel_file_path = f'{output_directory}\\VMVulnerabilityScan.{current_date}.xlsx'

try:
    workbook.save(excel_file_path)
    logging.info(f'Vulnerability data saved to {excel_file_path}')
except Exception as e:
    logging.error(f"Failed to write vulnerabilities to Excel file: {e}")
